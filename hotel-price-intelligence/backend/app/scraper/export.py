"""Xuất dữ liệu 1 crawl_run ra Excel — phục vụ gộp nhiều lần cào thành dataset huấn luyện sau này."""
import io
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

_COLUMNS = [
    ("hotel_id", "Hotel ID (slug)"),
    ("hotel_name", "Tên khách sạn"),
    ("city", "Thành phố"),
    ("address", "Địa chỉ"),
    ("review_score", "Điểm review"),
    ("review_count", "Số review"),
    ("crawl_url", "URL cào thực tế"),
    ("dom_room_row_count", "Số dòng DOM"),
    ("candidate_rate_count", "Candidate rate"),
    ("parsed_options_count", "Parser đọc được"),
    ("rejected_options_count", "Parser loại"),
    ("duplicate_options_count", "Option trùng đã bỏ"),
    ("raw_options_count", "Số option parser tạo"),
    ("saved_options_count", "Số option DB lưu"),
    ("observed_at", "Thời điểm cào (UTC)"),
    ("checkin_date", "Checkin"),
    ("checkout_date", "Checkout"),
    ("lead_time", "Lead time (ngày)"),
    ("room_type_raw", "Loại phòng (gốc)"),
    ("room_type_norm", "Loại phòng (chuẩn hoá)"),
    ("room_option_index", "Thứ tự option"),
    ("room_option_key", "Room option key"),
    ("room_identity_key", "Room identity key"),
    ("rate_plan_key", "Rate plan key"),
    ("is_reference_room", "Phòng tham chiếu"),
    ("reference_definition_id", "Reference definition ID"),
    ("reference_match_status", "Reference match"),
    ("reference_match_score", "Reference score"),
    ("price_total", "Giá tổng (VND)"),
    ("price_per_night", "Giá/đêm (VND)"),
    ("original_price", "Giá gốc (VND)"),
    ("discount_percent", "% giảm"),
    ("taxes_fees", "Thuế/phí tách riêng (VND)"),
    ("price_includes_tax", "Giá đã gồm thuế/phí"),
    ("max_occupancy", "Số khách tối đa"),
    ("bed_config", "Giường"),
    ("room_area", "Diện tích"),
    ("breakfast_included", "Bao gồm bữa sáng"),
    ("free_cancellation", "Huỷ miễn phí"),
    ("cancellation_policy", "Chính sách huỷ"),
    ("rooms_left", "Số phòng còn lại"),
    ("is_sold_out", "Hết phòng"),
    ("availability_status", "Trạng thái"),
    ("is_anomaly", "Bất thường"),
]

_ISSUE_COLUMNS = [
    ("id", "Item ID"),
    ("hotel_name_hint", "Tên từ file nguồn"),
    ("hotel_name", "Tên crawler nhận diện"),
    ("checkin_date", "Checkin"),
    ("status", "Trạng thái"),
    ("raw_options_count", "Số option parser tạo"),
    ("saved_options_count", "Số option DB lưu"),
    ("hotel_link", "URL cào thực tế"),
    ("error_message", "Lỗi / ghi chú"),
    ("dom_room_row_count", "Số dòng DOM"),
    ("candidate_rate_count", "Candidate"),
    ("parsed_options_count", "Parsed"),
    ("rejected_options_count", "Rejected"),
    ("duplicate_options_count", "Duplicate removed"),
    ("reference_match_status", "Reference match"),
    ("last_error_code", "Mã lỗi"),
    ("item_total_ms", "Tổng thời gian (ms)"),
]


def build_run_export_xlsx(
    run_id: int,
    rows: List[Dict[str, Any]],
    items: List[Dict[str, Any]] | None = None,
) -> bytes:
    """rows: list dict đã JOIN price_observations + hotels cho 1 crawl_run_id (xem repository)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"crawl_run_{run_id}"

    for col_idx, (_, header) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = openpyxl.styles.Alignment(vertical="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(_COLUMNS, start=1):
            value = row.get(key)
            if value is not None and key in ('observed_at', 'checkin_date', 'checkout_date'):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, (key, header) in enumerate(_COLUMNS, start=1):
        width = max(len(header) + 2, 14)
        if key in ('address', 'crawl_url', 'cancellation_policy'):
            width = 42
        elif key in ('room_type_raw', 'room_option_key'):
            width = 34
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 48)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24

    issue_items = [
        item for item in (items or [])
        if item.get('status') in ('error', 'partial', 'not_bookable')
    ]
    if issue_items:
        issue_ws = wb.create_sheet("item_issues")
        for col_idx, (_, header) in enumerate(_ISSUE_COLUMNS, start=1):
            cell = issue_ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="B45309")
            cell.alignment = openpyxl.styles.Alignment(vertical="center")

        for row_idx, item in enumerate(issue_items, start=2):
            for col_idx, (key, _) in enumerate(_ISSUE_COLUMNS, start=1):
                value = item.get(key)
                if value is not None and key == 'checkin_date':
                    value = str(value)
                issue_ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx, (key, header) in enumerate(_ISSUE_COLUMNS, start=1):
            width = max(len(header) + 2, 14)
            if key in ('hotel_link', 'error_message'):
                width = 52
            elif key in ('hotel_name_hint', 'hotel_name'):
                width = 32
            issue_ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 56)

        issue_ws.freeze_panes = "A2"
        issue_ws.auto_filter.ref = issue_ws.dimensions
        issue_ws.sheet_view.showGridLines = False
        issue_ws.row_dimensions[1].height = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
