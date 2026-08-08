"""Excel preflight/parser và compatibility entrypoint cho durable worker."""
from typing import List, Optional, Tuple

import openpyxl

from app.scraper.transform import _canonical_city
from app.scraper.url_utils import clean_hotel_link


def parse_hotel_list_excel(file_path: str) -> List[Tuple[str, Optional[str], Optional[str]]]:
    """Đọc nhiều sheet, cột A=tên và B=Booking URL; loại duplicate canonical URL."""
    links: List[Tuple[str, Optional[str], Optional[str]]] = []
    seen_links = set()
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell_a, cell_b in ws.iter_rows(min_row=2, max_col=2):
                if not cell_b.value:
                    continue
                link = str(cell_b.value).strip()
                name_hint = str(cell_a.value).strip() if cell_a.value else None
                if 'booking.com/hotel/' not in link:
                    continue
                canonical = clean_hotel_link(link).lower()
                if canonical in seen_links:
                    continue
                seen_links.add(canonical)
                links.append((link, name_hint, sheet_name))
    finally:
        wb.close()
    return links


def inspect_hotel_list_excel(file_path: str) -> dict:
    """Preflight workbook trước khi tạo job."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheets = []
    invalid_rows = []
    duplicate_rows = []
    seen_links = {}
    valid_links = 0
    total_rows = 0
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            city = _canonical_city(sheet_name)
            sheet_total = 0
            sheet_valid = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=2), start=2):
                name, value = row[0].value, row[1].value
                if value is None or not str(value).strip():
                    continue
                total_rows += 1
                sheet_total += 1
                link = str(value).strip()
                if 'booking.com/hotel/' not in link:
                    invalid_rows.append({
                        'sheet': sheet_name, 'row': row_idx, 'name': str(name or ''),
                        'reason': 'Không phải link khách sạn Booking.com',
                    })
                    continue
                canonical = clean_hotel_link(link).lower()
                if canonical in seen_links:
                    duplicate_rows.append({
                        'sheet': sheet_name, 'row': row_idx, 'name': str(name or ''),
                        'duplicate_of': seen_links[canonical],
                    })
                    continue
                seen_links[canonical] = f'{sheet_name}!B{row_idx}'
                valid_links += 1
                sheet_valid += 1
            sheets.append({
                'name': sheet_name, 'city': city, 'in_scope': city is not None,
                'total_rows': sheet_total, 'valid_links': sheet_valid,
            })
    finally:
        wb.close()
    return {
        'total_rows': total_rows,
        'valid_links': valid_links,
        'invalid_rows': invalid_rows,
        'duplicate_rows': duplicate_rows,
        'sheets': sheets,
        'search_context': '2 người lớn · 0 trẻ em · 1 phòng · 1 đêm · VND · anonymous',
    }


def drain_queue() -> None:
    """Compatibility one-shot: xử lý các item đã queued rồi thoát; không tự tạo scheduled job."""
    from app.scraper.worker import CrawlWorker
    CrawlWorker().run_until_empty()
