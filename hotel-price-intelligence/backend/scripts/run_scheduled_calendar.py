"""Generic calendar validator and durable scheduled-crawl runner."""
from __future__ import annotations

import argparse, hashlib, os, shutil, subprocess, sys, time
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.core.config import settings  # noqa: E402
from app.scraper.job_runner import inspect_hotel_list_excel, parse_hotel_list_excel  # noqa: E402

DEFAULT_PLAN_SHEET, DEFAULT_LOG_SHEET = "DAILY_CRAWL_PLAN", "CRAWL_LOG"
DEFAULT_CHECKIN_HEADERS = ["N1 gần", "N2 gần", "N3 gần", "N4 gần", "N5 gần", "N6 gần", "N7 gần", "N8 gần", "N9 gần", "F1 xa", "F2 xa", "F3 xa"]
AUX_GATES = ("Số ngày hợp lệ", "Không trùng local", "Không trùng VPS", "Không trùng nội bộ")
TERMINAL = {"completed", "failed"}


def _as_date(v: Any) -> date | None:
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v.strip(), fmt).date()
            except ValueError: pass
    return None


def _headers(ws) -> dict[str, int]:
    return {str(c.value).strip(): c.column for c in ws[3] if c.value is not None}


def _row(ws, headers: dict[str, int], target: date) -> int:
    if "Crawl date" not in headers: raise ValueError(f"{ws.title} thiếu Crawl date")
    rows = [r for r in range(4, ws.max_row + 1) if _as_date(ws.cell(r, headers["Crawl date"]).value) == target]
    if len(rows) != 1: raise ValueError(f"{ws.title} phải có đúng một dòng {target}, hiện có {len(rows)}")
    return rows[0]


def _writable(path: Path) -> None:
    try:
        with path.open("r+b"): pass
    except OSError as exc: raise RuntimeError(f"Workbook đang khóa/không ghi được: {path}") from exc


def _save(wb, path: Path) -> None:
    temp = path.with_name(f".{path.stem}.{os.getpid()}.tmp.xlsx")
    try:
        wb.save(temp)
        wb.close()
        os.replace(temp, path)
    finally: temp.unlink(missing_ok=True)


def load_contract(path: Path, target: date, plan_name: str, log_name: str, check_headers: list[str], environment: str, writable=False):
    if writable: _writable(path)
    wb = load_workbook(path)
    values_wb = load_workbook(path, data_only=True, read_only=True)
    if plan_name not in wb.sheetnames or log_name not in wb.sheetnames: raise ValueError("Thiếu plan/log sheet")
    plan, log = wb[plan_name], wb[log_name]
    value_plan = values_wb[plan_name]
    ph, lh = _headers(plan), _headers(log)
    pr, lr = _row(plan, ph, target), _row(log, lh, target)
    missing = [h for h in check_headers if h not in ph]
    if missing: raise ValueError(f"Thiếu check-in headers: {missing}")
    dates = [_as_date(value_plan.cell(pr, ph[h]).value) for h in check_headers]
    # openpyxl preserves formulas but cannot preserve Excel's cached results after
    # a save. Recompute AUX slots from the workbook's explicit SLOT_RULES contract.
    if environment == "local_aux" and any(v is None for v in dates):
        rules = wb["SLOT_RULES"]
        by_slot = {str(rules.cell(r, 1).value).strip(): (rules.cell(r, 4).value, int(rules.cell(r, 5).value))
                   for r in range(4, rules.max_row + 1)
                   if rules.cell(r, 1).value and rules.cell(r, 5).value is not None}
        dates = []
        for header in check_headers:
            anchor_value, cycle = by_slot[header]
            anchor = _as_date(anchor_value)
            if anchor is None: raise ValueError(f"SLOT_RULES thiếu anchor cho {header}")
            periods = 0 if target <= anchor else ((target - anchor).days - 1) // cycle + 1
            from datetime import timedelta
            dates.append(anchor + timedelta(days=periods * cycle))
    if any(v is None or v < target for v in dates): raise ValueError("Check-in rỗng/sai kiểu/trước crawl date")
    checkins = [v.isoformat() for v in dates if v]
    if len(set(checkins)) != len(check_headers): raise ValueError("Check-in bị trùng")
    if environment == "local_aux":
        local, vps = wb["LOCAL_REFERENCE"], wb["VPS_REFERENCE"]
        local_row, vps_row = _row(local, _headers(local), target), _row(vps, _headers(vps), target)
        local_dates = {_as_date(local.cell(local_row, c).value) for c in range(4, 16)}
        vps_dates = {_as_date(vps.cell(vps_row, c).value) for c in range(4, 9)}
        computed_gates = {
            "Số ngày hợp lệ": len(dates) == len(check_headers) and all(v is not None and v >= target for v in dates),
            "Không trùng local": not (set(dates) & local_dates),
            "Không trùng VPS": not (set(dates) & vps_dates),
            "Không trùng nội bộ": len(set(dates)) == len(dates),
        }
        for gate in AUX_GATES:
            cached = str(value_plan.cell(pr, ph[gate]).value or "").strip().upper() if gate in ph else ""
            if cached not in {"", "PASS"} or not computed_gates[gate]: raise ValueError(f"Gate {gate} chưa PASS")
    values_wb.close()
    return wb, log, lh, lr, checkins


def inspect_source(path: Path, expected: int):
    result, links = inspect_hotel_list_excel(str(path)), parse_hotel_list_excel(str(path))
    if int(result.get("valid_links", 0)) != expected or len(links) != expected:
        raise ValueError(f"Hotel file phải có đúng {expected} Booking links hợp lệ")
    for key in ("invalid_rows", "duplicate_rows", "out_of_scope_rows"):
        if result.get(key): raise ValueError(f"Hotel file có {key}: {result[key]}")
    return links


def validate(args, target: date):
    _, _, _, _, checkins = load_contract(args.calendar, target, args.plan_sheet, args.log_sheet, args.checkin_headers, args.environment)
    links = inspect_source(args.hotel_file, args.expected_hotels)
    if len(links) * len(checkins) != args.expected_items:
        raise ValueError(f"Expected items sai: {len(links)}×{len(checkins)} != {args.expected_items}")
    return checkins, links


def _write(log, headers, row, values):
    missing = [h for h in values if h not in headers]
    if missing: raise ValueError(f"CRAWL_LOG thiếu cột: {missing}")
    for h, v in values.items(): log.cell(row, headers[h]).value = v


def _ensure_worker(queue, backend: Path):
    if queue.worker_health().get("online"): return
    subprocess.Popen([sys.executable, str(backend / "scripts" / "run_worker.py")], cwd=backend, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0)
    for _ in range(30):
        time.sleep(2)
        if queue.worker_health().get("online"): return
    raise RuntimeError("Worker local không online sau 60 giây")


def _valid_records(run_id: int) -> int:
    from app.core.database import get_db_connection
    with get_db_connection() as conn:
        cur = conn.cursor(); cur.execute("SELECT COUNT(*) FROM price_observations WHERE crawl_run_id=%s", (run_id,))
        value = int(cur.fetchone()[0] or 0); cur.close(); return value


def run(args, target: date) -> int:
    from app.database.durable import DurableQueueRepository
    from app.database.repositories import CrawlRunRepository
    from app.scraper.data_contract import current_git_commit, default_crawl_context
    if args.resume_run_id:
        _, _, _, _, checkins = load_contract(args.calendar, target, args.plan_sheet, args.log_sheet, args.checkin_headers, args.environment)
        links = []
    else:
        checkins, links = validate(args, target)
    _writable(args.calendar)
    wb, log, lh, lr, _ = load_contract(args.calendar, target, args.plan_sheet, args.log_sheet, args.checkin_headers, args.environment, True)
    status, existing = str(log.cell(lr, lh["Status"]).value or "").strip(), log.cell(lr, lh["Run IDs"]).value
    if status in {"Hoàn thành", "Hoàn thành có lỗi"}: print(f"{target} đã terminal; không tạo run"); return 0
    queue, repo, backend = DurableQueueRepository(), CrawlRunRepository(), Path(__file__).resolve().parents[1]
    run_id = int(str(existing).split(",")[0].strip()) if existing else args.resume_run_id
    if run_id is None:
        if status != "Chưa chạy": raise ValueError(f"Chỉ tạo run khi Status='Chưa chạy', hiện là {status!r}")
        digest = hashlib.sha256(args.hotel_file.read_bytes()).hexdigest(); upload = (backend / settings.UPLOAD_DIR).resolve(); upload.mkdir(parents=True, exist_ok=True)
        saved = upload / f"{digest[:16]}_{args.hotel_file.name}"
        if not saved.exists(): shutil.copy2(args.hotel_file, saved)
        _ensure_worker(queue, backend); context = default_crawl_context(False); context["environment"] = args.environment
        run_id = queue.create_run_with_items(trigger_type="scheduled", source_file=str(saved), source_original_filename=args.hotel_file.name, source_file_sha256=digest, source_file_size=args.hotel_file.stat().st_size, date_mode="explicit", checkin_dates=checkins, hotel_links=links, crawl_context=context, save_artifacts=False, scraper_version=settings.SCRAPER_VERSION, selector_version=settings.SELECTOR_VERSION, git_commit=current_git_commit())
        created = repo.get_by_id(run_id)
        if not created or int(created.get("total") or 0) != args.expected_items: raise RuntimeError(f"Run {run_id} không đủ {args.expected_items} items")
        _write(log, lh, lr, {"Status":"Đang chạy", "Run IDs":str(run_id), "Started at":datetime.now(ZoneInfo(settings.DISPLAY_TIMEZONE)).replace(tzinfo=None), "Check-in count":len(checkins), "Environment":args.environment, "Notes":"Scheduled durable run; save_artifacts=false"}); _save(wb, args.calendar)
    elif not existing:
        # Explicit recovery after a run was durably created but the first workbook save failed.
        recovered = repo.get_by_id(run_id)
        if not recovered or recovered.get("trigger_type") != "scheduled" or int(recovered.get("total") or 0) != args.expected_items:
            raise ValueError(f"Không thể recovery scheduled run {run_id}")
        _write(log, lh, lr, {"Status":"Đang chạy", "Run IDs":str(run_id), "Started at":recovered.get("created_at") or datetime.now(ZoneInfo(settings.DISPLAY_TIMEZONE)).replace(tzinfo=None), "Check-in count":len(checkins), "Environment":args.environment, "Notes":f"Recovered monitor cho scheduled run {run_id}; không tạo run thay thế"}); _save(wb, args.calendar)
    while True:
        result = repo.get_by_id(run_id)
        if not result: raise RuntimeError(f"Không tìm thấy crawl run {run_id}")
        if result.get("status") in TERMINAL: break
        time.sleep(max(10, args.poll_seconds))
    wb, log, lh, lr, _ = load_contract(args.calendar, target, args.plan_sheet, args.log_sheet, args.checkin_headers, args.environment, True)
    partial, errors = int(result.get("partial_count") or 0), int(result.get("error_count") or 0)
    final = "Hoàn thành có lỗi" if result.get("status") == "failed" or partial or errors else "Hoàn thành"
    _write(log, lh, lr, {"Status":final,"Run IDs":str(run_id),"Finished at":datetime.now(ZoneInfo(settings.DISPLAY_TIMEZONE)).replace(tzinfo=None),"Total items":int(result.get("total") or 0),"Processed":int(result.get("processed") or 0),"Success":int(result.get("success_count") or 0),"Partial":partial,"Sold out":int(result.get("sold_out_count") or 0),"Not bookable":int(result.get("not_bookable_count") or 0),"Error":errors,"Valid records":_valid_records(run_id),"Check-in count":len(checkins),"Environment":args.environment,"Notes":f"Scheduled run {run_id} terminal"}); _save(wb, args.calendar)
    return 0


def build_parser():
    p = argparse.ArgumentParser(); p.add_argument("--calendar", type=Path, required=True); p.add_argument("--hotel-file", type=Path, required=True); p.add_argument("--date"); p.add_argument("--poll-seconds", type=int, default=60)
    p.add_argument("--plan-sheet", default=DEFAULT_PLAN_SHEET); p.add_argument("--log-sheet", default=DEFAULT_LOG_SHEET); p.add_argument("--checkin-headers", default=",".join(DEFAULT_CHECKIN_HEADERS)); p.add_argument("--environment", default="local_primary"); p.add_argument("--expected-hotels", type=int, default=355); p.add_argument("--expected-items", type=int, default=4260); p.add_argument("--resume-run-id", type=int); p.add_argument("--validate-only", action="store_true"); return p


def main() -> int:
    args = build_parser().parse_args(); args.calendar = args.calendar.resolve(); args.hotel_file = args.hotel_file.resolve(); args.checkin_headers = [x.strip() for x in args.checkin_headers.split(",") if x.strip()]
    backend = Path(__file__).resolve().parents[1]
    for path in (args.calendar, args.hotel_file, Path(__file__), backend / ".env"):
        if not path.exists(): raise FileNotFoundError(path)
    target = datetime.strptime(args.date, "%Y-%m-%d").date() if args.date else datetime.now(ZoneInfo(settings.DISPLAY_TIMEZONE)).date()
    if args.validate_only:
        checkins, links = validate(args, target); print(f"VALIDATE-ONLY PASS date={target} hotels={len(links)} checkins={len(checkins)} items={len(links)*len(checkins)} environment={args.environment}"); return 0
    try: return run(args, target)
    except Exception as exc:
        try:
            wb, log, lh, lr, _ = load_contract(args.calendar, target, args.plan_sheet, args.log_sheet, args.checkin_headers, args.environment, True); existing = log.cell(lr, lh["Run IDs"]).value
            _write(log, lh, lr, {"Status":"Đang chạy" if existing else "Thất bại", "Notes":f"Run {existing} đã tồn tại; cần tiếp tục kiểm tra: {type(exc).__name__}: {exc}" if existing else f"{type(exc).__name__}: {exc}"}); _save(wb, args.calendar)
        except Exception: pass
        raise


if __name__ == "__main__": raise SystemExit(main())
