import io
import zipfile
from datetime import datetime
from urllib.parse import parse_qs, urlparse

import openpyxl

from app.scraper.artifacts import save_page_artifacts
from app.scraper.booking_scraper import _extract_price, _extract_rooms, _extract_tax_info, _get_room_name
from app.scraper.export import build_run_export_xlsx
from app.scraper.job_runner import inspect_hotel_list_excel, parse_hotel_list_excel
from app.scraper.parser import infer_max_occupancy
from app.scraper.reference import rate_plan_key, room_identity_key, select_best_match
from app.scraper.transform import build_price_observations
from app.scraper.url_utils import build_scrape_url, set_checkin_checkout


def test_date_rewrite_keeps_valid_query_when_checkin_is_first_param():
    result = set_checkin_checkout(
        'https://www.booking.com/hotel/vn/example.html?checkin=2025-01-01&checkout=2025-01-02',
        '2026-08-16',
        '2026-08-17',
    )
    query = parse_qs(urlparse(result).query)
    assert query['checkin'] == ['2026-08-16']
    assert query['checkout'] == ['2026-08-17']


def test_scrape_url_normalizes_occupancy_currency_and_language():
    result = build_scrape_url(
        'https://www.booking.com/hotel/vn/example.html?no_rooms=4&group_adults=8&room1=A&room2=A&age=0&req_age=0',
        '2026-08-16',
        '2026-08-17',
    )
    query = parse_qs(urlparse(result).query)
    assert query['no_rooms'] == ['1']
    assert query['group_adults'] == ['2']
    assert query['group_children'] == ['0']
    assert query['room1'] == ['A,A']
    assert 'room2' not in query
    assert 'age' not in query
    assert 'req_age' not in query
    assert query['selected_currency'] == ['VND']
    assert query['lang'] == ['vi']


def test_scrape_url_drops_stale_tracking_and_session_parameters():
    result = build_scrape_url(
        'https://www.booking.com/hotel/vn/example.vi.html?aid=1&label=old&sid=stale&srpvid=old',
        '2026-08-29',
        '2026-08-30',
    )
    query = parse_qs(urlparse(result).query)

    assert 'aid' not in query
    assert 'label' not in query
    assert 'sid' not in query
    assert 'srpvid' not in query
    assert query['checkin'] == ['2026-08-29']


def test_excel_reader_does_not_require_dimension_metadata(tmp_path):
    source = tmp_path / 'source.xlsx'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Phú Quốc'
    sheet.append(['Tên khách sạn', 'Link'])
    sheet.append(['Example', 'https://www.booking.com/hotel/vn/example.vi.html'])
    workbook.save(source)

    without_dimension = tmp_path / 'without_dimension.xlsx'
    with zipfile.ZipFile(source, 'r') as input_zip, zipfile.ZipFile(without_dimension, 'w') as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == 'xl/worksheets/sheet1.xml':
                content = content.replace(b'<dimension ref="A1:B2"/>', b'')
            output_zip.writestr(entry, content)

    preflight = inspect_hotel_list_excel(str(without_dimension))
    parsed = parse_hotel_list_excel(str(without_dimension))
    assert preflight['valid_links'] == 1
    assert preflight['sheets'][0]['city'] == 'Phú Quốc'
    assert len(parsed) == 1


def test_room_options_are_not_deduplicated_by_room_type_norm():
    raw = {
        'rooms': [
            {
                'room_type_raw': 'Phòng Standard',
                'max_occupancy': 2,
                'price_per_night': 1_000_000,
                'original_price': None,
                'discount_percent': None,
                'bed_options': '1 giường đôi',
                'room_area': '20 m²',
                'price_includes_tax': True,
                'taxes_fees': None,
                'facility_lines': ['Không hoàn tiền'],
            },
            {
                'room_type_raw': 'Phòng Standard',
                'max_occupancy': 2,
                'price_per_night': 1_100_000,
                'original_price': None,
                'discount_percent': None,
                'bed_options': '1 giường đôi',
                'room_area': '20 m²',
                'price_includes_tax': True,
                'taxes_fees': None,
                'facility_lines': ['Miễn phí hủy'],
            },
        ]
    }
    records = build_price_observations(
        raw, 'example', 1, 'manual', datetime(2026, 8, 5, 12, 0),
        '2026-08-16', '2026-08-17',
    )
    assert len(records) == 2
    assert [record['room_option_index'] for record in records] == [0, 1]
    assert len({record['room_option_key'] for record in records}) == 2
    assert sum(record['is_reference_room'] for record in records) == 1


def test_infer_max_occupancy_prefers_explicit_room_name():
    assert infer_max_occupancy('Phòng 3 Người Nhìn Ra Thành Phố', 2) == 3
    assert infer_max_occupancy('Comfort Quadruple Room', 2) == 4
    assert infer_max_occupancy('Phòng Deluxe', 2) == 2


class _TaxRow:
    text = ''

    def __init__(self, text):
        self._text = text

    def get_attribute(self, name):
        return self._text if name == 'textContent' else None


def test_tax_parser_does_not_invent_tax_amount():
    amount, includes = _extract_tax_info(_TaxRow('VND 1.500.000 Đã bao gồm thuế và phí'))
    assert amount is None
    assert includes is True


class _HiddenElement:
    text = ''

    def __init__(self, text_content):
        self.text_content = text_content

    def get_attribute(self, name):
        return self.text_content if name == 'textContent' else None


class _HiddenContainer:
    def __init__(self, child):
        self.child = child

    def find_element(self, by, selector):
        return self.child


def test_hidden_booking_rate_uses_dom_text_content():
    room_header = _HiddenContainer(_HiddenElement('  Phòng Giường Đôi Hạng Tiết Kiệm\n'))
    price_row = _HiddenContainer(_HiddenElement('\nVND\u00a0550.000\n'))

    assert _get_room_name(room_header) == 'Phòng Giường Đôi Hạng Tiết Kiệm'
    assert _extract_price(price_row, ['.bui-price-display__value']) == 550_000


def test_export_contains_tax_and_audit_columns():
    content = build_run_export_xlsx(7, [{
        'hotel_id': 'example',
        'hotel_name': 'Example Hotel',
        'room_option_index': 0,
        'room_option_key': 'a' * 64,
        'taxes_fees': None,
        'price_includes_tax': True,
        'raw_options_count': 1,
        'saved_options_count': 1,
    }])
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    assert 'Thuế/phí tách riêng (VND)' in headers
    assert 'Giá đã gồm thuế/phí' in headers
    assert 'URL cào thực tế' in headers
    assert sheet.auto_filter.ref == sheet.dimensions


def test_export_includes_separate_issue_sheet():
    content = build_run_export_xlsx(8, [], [{
        'id': 3,
        'hotel_name_hint': 'Dead Hotel',
        'hotel_name': None,
        'checkin_date': '2026-08-17',
        'status': 'error',
        'raw_options_count': 0,
        'saved_options_count': 0,
        'hotel_link': 'https://www.booking.com/hotel/vn/dead.html',
        'error_message': 'Link chỗ nghỉ không còn mở được',
    }])
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    assert workbook.sheetnames == ['crawl_run_8', 'item_issues']
    issue_sheet = workbook['item_issues']
    assert issue_sheet['B2'].value == 'Dead Hotel'
    assert issue_sheet['E2'].value == 'error'
    assert issue_sheet['I2'].value == 'Link chỗ nghỉ không còn mở được'


class _MissingHeaderRow:
    def find_elements(self, by, selector):
        return []


class _SingleRowDriver:
    def find_elements(self, by, selector):
        return [_MissingHeaderRow()] if selector == 'tr.js-rt-block-row' else []


def test_parser_completeness_accounts_for_every_candidate():
    result = {'rooms': []}
    diagnostics = _extract_rooms(_SingleRowDriver(), result)

    assert diagnostics['candidate_rate_count'] == 1
    assert diagnostics['parsed_options_count'] == 0
    assert diagnostics['rejected_options_count'] == 1
    assert diagnostics['rejected_options'][0]['reason_code'] == 'missing_room_header'


def test_reference_match_uses_stable_room_and_rate_plan_identity():
    room = {
        'room_type_raw': 'Phòng Deluxe Hướng Biển',
        'max_occupancy': 2,
        'bed_config': '1 giường đôi',
        'room_area': '28 m²',
        'breakfast_included': True,
        'free_cancellation': True,
        'cancellation_policy': 'Miễn phí hủy',
    }
    room['room_identity_key'] = room_identity_key(room)
    room['rate_plan_key'] = rate_plan_key(room)
    reference = {
        **room,
        'room_type_anchor_raw': room['room_type_raw'],
    }

    assert select_best_match([room], reference) == (0, 'exact', 1.0)


class _ArtifactDriver:
    page_source = '<html><body>snapshot</body></html>'

    def save_screenshot(self, path):
        with open(path, 'wb') as output:
            output.write(b'png')
        return True


def test_artifact_writer_saves_html_gzip_and_screenshot_when_invoked(tmp_path):
    artifacts = save_page_artifacts(_ArtifactDriver(), str(tmp_path), 11, 22)

    assert artifacts['artifact_html_path'].endswith('11\\22\\page.html.gz')
    assert artifacts['screenshot_path'].endswith('11\\22\\page.png')
