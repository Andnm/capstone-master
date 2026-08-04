import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
from io import BytesIO, StringIO
import time
import openpyxl
import json
from app.services.booking_scraper import get_driver

try:
    from selenium import webdriver
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False


def is_agoda_link(text):
    if pd.isna(text) or text is None:
        return False
    text_str = str(text).strip()
    return 'agoda.com' in text_str.lower() and ('/hotel/' in text_str or '/book/' in text_str or 'hotelid=' in text_str.lower())


def extract_hyperlinks_from_excel(file_bytes):
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active
        links_info = []

        for col_idx in [1, 2, 3]:
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.hyperlink and cell.hyperlink.target:
                    link = cell.hyperlink.target
                    is_valid = is_agoda_link(link)
                    links_info.append({
                        'row': row_idx,
                        'col': chr(64 + col_idx),
                        'link': link.strip(),
                        'cell_value': str(cell.value) if cell.value else '',
                        'is_valid': is_valid,
                        'note': '' if is_valid else '⚠️Link không hợp lệ'
                    })

                elif cell.value:
                    cell_text = str(cell.value).strip()
                    if 'http' in cell_text.lower() or 'www.' in cell_text.lower():
                        is_valid = is_agoda_link(cell_text)
                        links_info.append({
                            'row': row_idx,
                            'col': chr(64 + col_idx),
                            'link': cell_text,
                            'cell_value': cell_text,
                            'is_valid': is_valid,
                            'note': '' if is_valid else '⚠️Link không hợp lệ'
                        })
        
        return links_info
    except Exception as e:
        print(f"Lỗi khi đọc hyperlink từ Excel: {str(e)}")
        return []


def find_agoda_links(df):
    links_info = []
    
    for col_idx in [0, 1, 2]:
        if col_idx < len(df.columns):
            col_name = df.columns[col_idx]
            for row_idx, value in enumerate(df[col_name]):
                if pd.isna(value) or value is None:
                    continue
                    
                value_str = str(value).strip()

                if value_str.startswith('=HYPERLINK('):
                    match = re.search(r'=HYPERLINK\("([^"]+)"', value_str)
                    if match:
                        url = match.group(1)
                        is_valid = is_agoda_link(url)
                        text_match = re.search(r',\s*"([^"]+)"\)', value_str)
                        display_text = text_match.group(1) if text_match else url
                        links_info.append({
                            'row': row_idx + 1,
                            'col': chr(65 + col_idx),
                            'link': url.strip(),
                            'cell_value': display_text,
                            'is_valid': is_valid,
                            'note': '' if is_valid else '⚠️Link không hợp lệ'
                        })

                elif 'http' in value_str.lower() or 'www.' in value_str.lower():
                    is_valid = is_agoda_link(value_str)
                    links_info.append({
                        'row': row_idx + 1,
                        'col': chr(65 + col_idx),
                        'link': value_str,
                        'cell_value': value_str,
                        'is_valid': is_valid,
                        'note': '' if is_valid else '⚠️Link không hợp lệ'
                    })
    
    return links_info


def load_google_sheet(url):
    try:
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not match:
            return None, None, "URL Google Sheets không hợp lệ"
        
        spreadsheet_id = match.group(1)

        gid_match = re.search(r'[#&]gid=([0-9]+)', url)
        gid = gid_match.group(1) if gid_match else '0'

        csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
        csv_response = requests.get(csv_url, timeout=10)
        csv_response.raise_for_status()
        df = pd.read_csv(StringIO(csv_response.text), header=None)

        links_info = []

        try:
            html_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=html&gid={gid}"
            html_response = requests.get(html_url, timeout=10)
            
            if html_response.status_code == 200:
                soup = BeautifulSoup(html_response.content, 'lxml')

                for row_idx, tr in enumerate(soup.find_all('tr'), 1):
                    cells = tr.find_all('td')
                    for col_idx in range(min(3, len(cells))):
                        cell = cells[col_idx]

                        link_tag = cell.find('a', href=True)
                        if link_tag:
                            href = link_tag.get('href')
                            is_valid = is_agoda_link(href)
                            links_info.append({
                                'row': row_idx,
                                'col': chr(65 + col_idx),
                                'link': href.strip(),
                                'cell_value': link_tag.get_text(strip=True),
                                'is_valid': is_valid,
                                'note': '' if is_valid else '⚠️Link không hợp lệ'
                            })
                        else:
                            cell_text = cell.get_text(strip=True)
                            if 'http' in cell_text.lower() or 'www.' in cell_text.lower():
                                is_valid = is_agoda_link(cell_text)
                                links_info.append({
                                    'row': row_idx,
                                    'col': chr(65 + col_idx),
                                    'link': cell_text,
                                    'cell_value': cell_text,
                                    'is_valid': is_valid,
                                    'note': '' if is_valid else '⚠️Link không hợp lệ'
                                })
            else:
                pubhtml_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/pubhtml?gid={gid}"
                html_response = requests.get(pubhtml_url, timeout=10)
                
                if html_response.status_code == 200:
                    soup = BeautifulSoup(html_response.content, 'lxml')
                    
                    for row_idx, tr in enumerate(soup.find_all('tr'), 1):
                        cells = tr.find_all('td')
                        for col_idx in range(min(3, len(cells))):
                            cell = cells[col_idx]
                            
                            link_tag = cell.find('a', href=True)
                            if link_tag:
                                href = link_tag.get('href')
                                is_valid = is_agoda_link(href)
                                links_info.append({
                                    'row': row_idx,
                                    'col': chr(65 + col_idx),
                                    'link': href.strip(),
                                    'cell_value': link_tag.get_text(strip=True),
                                    'is_valid': is_valid,
                                    'note': '' if is_valid else '⚠️Link không hợp lệ'
                                })
                            else:
                                cell_text = cell.get_text(strip=True)
                                if 'http' in cell_text.lower() or 'www.' in cell_text.lower():
                                    is_valid = is_agoda_link(cell_text)
                                    links_info.append({
                                        'row': row_idx,
                                        'col': chr(65 + col_idx),
                                        'link': cell_text,
                                        'cell_value': cell_text,
                                        'is_valid': is_valid,
                                        'note': '' if is_valid else '⚠️Link không hợp lệ'
                                    })
        except Exception as e:
            print(f"HTML extraction error: {e}")

        if not links_info:
            for col_idx in [0, 1, 2]:
                if col_idx < len(df.columns):
                    col_name = df.columns[col_idx]
                    for row_idx, value in enumerate(df[col_name]):
                        if pd.notna(value):
                            value_str = str(value).strip()
                            if 'http' in value_str.lower() or 'www.' in value_str.lower():
                                is_valid = is_agoda_link(value_str)
                                links_info.append({
                                    'row': row_idx + 1,
                                    'col': chr(65 + col_idx),
                                    'link': value_str,
                                    'cell_value': value_str,
                                    'is_valid': is_valid,
                                    'note': '' if is_valid else '⚠️Link không hợp lệ'
                                })
        
        return df, links_info, None
    except Exception as e:
        return None, None, f"Lỗi khi tải Google Sheets: {str(e)}"


def scrape_agoda_data(url):
    if not SELENIUM_AVAILABLE:
        return None, "Selenium chưa được cài đặt. Vui lòng chạy: pip install selenium webdriver-manager"
    
    import os
    
    try:
        os.system('taskkill /F /IM msedge.exe /T >nul 2>&1')
        os.system('taskkill /F /IM msedgedriver.exe /T >nul 2>&1')
        time.sleep(2)
    except:
        pass
    
    driver = None

    try:
        driver = get_driver(is_headless=True)
        driver.set_page_load_timeout(45)
        
        # Retry logic for page load (Phase 1: Immediate Retry)
        
        # Retry logic for page load (Phase 1: Immediate Retry)
        max_retries = 3
        for attempt in range(max_retries):
            try:
                driver.get(url)
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.TAG_NAME, 'body'))
                )
                break
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(5)
        
        time.sleep(10)

        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))
            )
        except:
            pass
        
        time.sleep(3)
        
        result = {
            'hotel_name': None,
            'rooms': [],
            'rating': None,
            'review_count': None
        }

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        
        hotel_name_elem = soup.select_one('h3[data-testid="property-name-id"]')
        if hotel_name_elem:
            result['hotel_name'] = hotel_name_elem.get_text(strip=True)
        if not result['hotel_name']:
            next_data_script = soup.find('script', {'id': '__NEXT_DATA__', 'type': 'application/json'})
            if next_data_script:
                try:
                    json_data = json.loads(next_data_script.string)
                    page_props = json_data.get('props', {}).get('pageProps', {})
                    for key in ['hotelName', 'propertyName', 'name', 'title']:
                        if key in page_props and page_props[key]:
                            result['hotel_name'] = page_props[key]
                            break
                except:
                    pass
        if not result['hotel_name']:
            h1_tags = soup.find_all('h1')
            for h1 in h1_tags:
                text = h1.get_text().strip()
                if text and 5 <= len(text) <= 200:
                    if not any(skip in text.lower() for skip in ['cookie', 'search', 'menu', 'login', 'đăng']):
                        result['hotel_name'] = text
                        break
        
        rating_elem = soup.select_one('span.sc-dmqHEX.cCFmSd')
        if rating_elem:
            text = rating_elem.get_text(strip=True)
            rating_match = re.search(r'(\d+[,.]?\d*)', text.replace(',', '.'))
            if rating_match:
                result['rating'] = rating_match.group(1)
        
        review_elem = soup.select_one('[data-testid="review-count"]')
        if review_elem:
            text = review_elem.get_text(strip=True)
            count_match = re.search(r'(\d+)', text.replace(',', '').replace('.', ''))
            if count_match:
                result['review_count'] = count_match.group(1)
        
        room_name_elem = soup.select_one('h4[data-testid="room-heading"]')
        room_name = room_name_elem.get_text(strip=True) if room_name_elem else None
        
        prices = {}
        
        orig_price_elem = soup.select_one('[data-element-name="fpc-cor-price"]')
        if orig_price_elem:
            prices['original_price'] = orig_price_elem.get('data-fpc-value', orig_price_elem.get_text(strip=True))
        
        our_price_elem = soup.select_one('[data-element-name="fpc-our-price"]')
        if our_price_elem:
            prices['our_price'] = our_price_elem.get('data-fpc-value', our_price_elem.get_text(strip=True))
        
        room_price_elem = soup.select_one('[data-element-name="fpc-room-price"]')
        if room_price_elem:
            prices['room_price'] = room_price_elem.get('data-fpc-value', room_price_elem.get_text(strip=True))
        
        final_price_elem = soup.select_one('[data-element-name="fpc-total-price"]')
        if final_price_elem:
            prices['final_price'] = final_price_elem.get('data-fpc-value', final_price_elem.get_text(strip=True))
        
        tax_elem = soup.select_one('[data-element-name="fpc-tax-and-fee-amount"]')
        if tax_elem:
            prices['tax_and_fee'] = tax_elem.get('data-fpc-value', tax_elem.get_text(strip=True))
        
        if room_name or prices:
            room_info = {
                'room_type': room_name,
                'price': prices.get('final_price', prices.get('room_price', 'N/A')),
                'price_original': prices.get('original_price'),
                'prices_detail': prices
            }
            result['rooms'].append(room_info)
        
        return result, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"Lỗi: {str(e)}"
        
    finally:
        if driver:
            driver.quit()
