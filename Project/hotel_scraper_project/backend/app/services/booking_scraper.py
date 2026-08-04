import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
from datetime import datetime
from io import BytesIO, StringIO
import time
import openpyxl
import json
import tempfile
import os
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse

import random

def get_random_user_agent():
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    ]
    return random.choice(user_agents)

try:
    from selenium import webdriver
    from selenium.webdriver.edge.service import Service as EdgeService
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

def get_driver(is_headless=True):
    # Check if running in Docker/Linux with Chrome installed
    is_docker = os.path.exists('/.dockerenv') or os.path.exists('/usr/bin/google-chrome')
    
    if is_docker:
        # Use Chrome in Docker
        options = ChromeOptions()
        options.add_argument('--headless=new')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--lang=vi-VN')
        options.add_argument(f'user-agent={get_random_user_agent()}')
        
        # Anti-detection arguments
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-extensions')
        options.add_argument('--no-first-run')
        options.add_argument('--disable-web-security')
        options.add_argument('--disable-features=VizDisplayCompositor')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # Location and language preferences
        prefs = {
            "profile.default_content_setting_values.geolocation": 1,
            "profile.managed_default_content_settings.geolocation": 1,
            "profile.content_settings.exceptions.geolocation": {
                "*": {"setting": 1}
            },
            "intl.accept_languages": "vi-VN,vi,en",
            "profile.default_content_settings.popups": 0
        }
        options.add_experimental_option("prefs", prefs)
        
        # Use webdriver_manager
        from webdriver_manager.chrome import ChromeDriverManager
        
        service = ChromeService(ChromeDriverManager().install())
        
        driver = webdriver.Chrome(service=service, options=options)
        
        # Remove automation indicators
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        # Setup location override (GPS spoofing only)
        try:
            # Set geolocation to Vietnam
            driver.execute_cdp_cmd("Emulation.setGeolocationOverride", {
                "latitude": 21.028511,
                "longitude": 105.854164,
                "accuracy": 100
            })
            
            # Set timezone to Vietnam
            driver.execute_cdp_cmd("Emulation.setTimezoneOverride", {
                "timezoneId": "Asia/Ho_Chi_Minh"
            })
            
            # Set locale to Vietnamese
            driver.execute_cdp_cmd("Emulation.setLocaleOverride", {
                "locale": "vi-VN"
            })
            
        except Exception as e:
            pass
            
        return driver
    else:
        # Use Chrome locally (Windows) with fallback to Edge
        driver = None
        last_error = None
        
        # Try Chrome first (more reliable CDN)
        try:
            from webdriver_manager.chrome import ChromeDriverManager
            if not is_headless:
                temp_dir = tempfile.mkdtemp(prefix='selenium_booking_')
            
            options = ChromeOptions()
            if is_headless:
                options.add_argument('--headless=new')
                
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--lang=vi-VN')
            options.add_argument(f'user-agent={get_random_user_agent()}')
            
            # Anti-detection arguments
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            options.add_argument('--disable-web-security')
            
            # Location and language preferences
            prefs = {
                "profile.default_content_setting_values.geolocation": 1,
                "profile.managed_default_content_settings.geolocation": 1,
                "intl.accept_languages": "vi-VN,vi,en",
                "profile.default_content_settings.popups": 0
            }
            options.add_experimental_option("prefs", prefs)
            
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            
            # Remove automation indicators
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            # Apply GPS spoofing
            try:
                driver.execute_cdp_cmd("Emulation.setGeolocationOverride", {
                    "latitude": 21.028511,
                    "longitude": 105.854164,
                    "accuracy": 100
                })
                
                driver.execute_cdp_cmd("Emulation.setTimezoneOverride", {
                    "timezoneId": "Asia/Ho_Chi_Minh"
                })
                
                driver.execute_cdp_cmd("Emulation.setLocaleOverride", {
                    "locale": "vi-VN"
                })
                
            except Exception as e:
                pass
            
            return driver
            
        except Exception as chrome_error:
            last_error = chrome_error
        
        # Fallback to Edge if Chrome fails
        try:
            from webdriver_manager.microsoft import EdgeChromiumDriverManager
            if not is_headless:
                temp_dir = tempfile.mkdtemp(prefix='selenium_booking_')
            
            options = EdgeOptions()
            options.use_chromium = True
            if is_headless:
                options.add_argument('--headless=new')
                
            options.add_argument('--disable-gpu')
            options.add_argument('--lang=vi-VN')
            options.add_argument(f'user-agent={get_random_user_agent()}')
            
            # Enhanced GPS spoofing for Edge - removed proxy settings
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            options.add_argument('--disable-web-security')
            options.add_argument('--allow-running-insecure-content')
            
            # Fake geolocation via prefs
            prefs = {
                "profile.default_content_setting_values.geolocation": 1,
                "profile.managed_default_content_settings.geolocation": 1,
                "profile.default_content_settings.geolocation": 1,
                "intl.accept_languages": "vi-VN,vi,en-US,en"
            }
            options.add_experimental_option("prefs", prefs)
            
            # Install EdgeDriver
            service = EdgeService(EdgeChromiumDriverManager().install())
            driver = webdriver.Edge(service=service, options=options)
            
            # Remove automation indicators
            driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            # Apply GPS spoofing for Edge
            try:
                driver.execute_cdp_cmd("Emulation.setGeolocationOverride", {
                    "latitude": 21.028511,
                    "longitude": 105.854164,
                    "accuracy": 100
                })
                
                driver.execute_cdp_cmd("Emulation.setTimezoneOverride", {
                    "timezoneId": "Asia/Ho_Chi_Minh"
                })
                
                driver.execute_cdp_cmd("Emulation.setLocaleOverride", {
                    "locale": "vi-VN"
                })
                
            except Exception as e:
                pass
            
            return driver
            
        except Exception as edge_error:
            raise Exception(f"Failed to initialize browser driver. Chrome error: {last_error}. Edge error: {edge_error}")

def is_booking_link(text):
    if pd.isna(text) or text is None:
        return False
    text_str = str(text).strip()
    return 'booking.com/hotel/' in text_str


def get_markets_from_excel(file_bytes):
    """Lấy danh sách các markets (sheet names) từ Excel file"""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        return wb.sheetnames
    except Exception as e:
        print(f"Lỗi khi đọc sheet names từ Excel: {str(e)}")
        return []


def extract_hyperlinks_from_excel(file_bytes, market=None):
    """
    Trích xuất links từ Excel file.
    Format: Cột A = Tên khách sạn, Cột B = Link
    Nếu market=None: lấy tất cả sheets với thông tin market
    Nếu market được chỉ định: chỉ lấy từ sheet đó
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        links_info = []
        
        # Xác định các sheets cần xử lý
        if market:
            # Nếu chỉ định market, chỉ xử lý sheet đó
            sheets_to_process = [(market, wb[market])] if market in wb.sheetnames else []
        else:
            # Nếu không chỉ định, xử lý tất cả sheets
            sheets_to_process = [(sheet_name, wb[sheet_name]) for sheet_name in wb.sheetnames]
        
        for sheet_name, ws in sheets_to_process:
            # Format: A = Tên KS, B = Link
            for row_idx in range(2, ws.max_row + 1):  # Bỏ qua header row
                cell_a = ws.cell(row=row_idx, column=1)  # Cột A - Tên khách sạn
                cell_b = ws.cell(row=row_idx, column=2)  # Cột B - Link

                # Lấy giá trị text trực tiếp từ cột B (Link)
                if cell_b.value:
                    cell_text = str(cell_b.value).strip()
                    if 'http' in cell_text.lower() or 'www.' in cell_text.lower():
                        is_valid = is_booking_link(cell_text)
                        
                        # Lấy tên khách sạn từ cột A
                        hotel_name = str(cell_a.value).strip() if cell_a.value else ''
                        
                        links_info.append({
                            'row': row_idx,
                            'col': 'B',
                            'link': cell_text,
                            'cell_value': cell_text,
                            'hotel_name': hotel_name,
                            'is_valid': is_valid,
                            'market': sheet_name,
                            'note': '' if is_valid else 'Link không hợp lệ'
                        })
        
        return links_info
    except Exception as e:
        print(f"Lỗi khi đọc hyperlink từ Excel: {str(e)}")
        return []


def find_booking_links(df):
    links_info = []
    
    # Chỉ lấy link từ cột B (index 1)
    if len(df.columns) > 1:
        col_b = df.columns[1]  # Cột B
        col_a = df.columns[0] if len(df.columns) > 0 else None  # Cột A (tên khách sạn)
        
        for row_idx, value in enumerate(df[col_b]):
            if pd.isna(value) or value is None:
                continue
                
            value_str = str(value).strip()
            
            # Lấy tên khách sạn từ cột A
            hotel_name = ''
            if col_a is not None:
                hotel_value = df[col_a].iloc[row_idx]
                if not pd.isna(hotel_value) and hotel_value is not None:
                    hotel_name = str(hotel_value).strip()

            # Chỉ lấy text value trực tiếp, không parse HYPERLINK formula
            if 'http' in value_str.lower() or 'www.' in value_str.lower():
                is_valid = is_booking_link(value_str)
                links_info.append({
                    'row': row_idx + 1,
                    'col': 'B',
                    'link': value_str,
                    'cell_value': value_str,
                    'hotel_name': hotel_name,
                    'is_valid': is_valid,
                    'note': '' if is_valid else '⚠️Link không hợp lệ'
                })
    
    return links_info


def force_vnd_currency(url):
    try:
        parsed = urlparse(str(url))
        query = parse_qs(parsed.query)
        query['selected_currency'] = ['VND']
        query['lang'] = ['vi']
        new_query = urlencode(query, doseq=True)
        return urlunparse(parsed._replace(query=new_query))
    except Exception:
        return url


def get_sheets_from_google_spreadsheet(url):
    """
    Lấy danh sách tất cả sheets từ Google Spreadsheet.
    Returns: dict {sheet_name: gid} hoặc None nếu lỗi
    """
    try:
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not match:
            return None
        
        spreadsheet_id = match.group(1)
        
        # Method 1: Try to get sheets from HTML page structure
        try:
            html_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
            response = requests.get(html_url, timeout=15)
            
            if response.status_code != 200:
                return None
            
            content = response.text
            sheets = {}
            
            # Try multiple patterns to find sheet info
            # Pattern 1: Look for sheet data in format: ["SheetName",number,...]
            # This appears in the _docs_flag_initialData variable
            patterns_to_try = [
                # Pattern: ["SheetName",gid, with constraints
                r'\["([^"]{2,50})",\s*(\d+)',
                # Pattern: Sheet tabs in HTML with data attributes  
                r'data-sheet-name="([^"]+)"[^>]*data-sheet-id="(\d+)"',
                # Pattern: In JavaScript object
                r'"name":"([^"]+)","sheetId":(\d+)',
            ]
            
            for pattern in patterns_to_try:
                matches = re.findall(pattern, content)
                if matches:
                    break
            
            if not matches:
                return None
            
            # Keywords to filter out (technical/system fields)
            exclude_keywords = [
                'docs.', 'http', 'www', '.com', 'security', 'access', 
                'capabilities', 'metadata', 'settings', 'config', 'api',
                'permissions', 'sharing', 'export', 'import', 'sync'
            ]
            
            # Lọc để chỉ lấy những cái có vẻ là sheet names
            seen_gids = set()
            valid_count = 0
            
            for sheet_name, gid in matches:
                # Skip if already seen
                if gid in seen_gids:
                    continue
                
                # Check if should exclude
                should_exclude = any(kw in sheet_name.lower() for kw in exclude_keywords)
                
                # Validate sheet name
                is_valid = (
                    sheet_name and 
                    not should_exclude and
                    not sheet_name.startswith('_') and 
                    2 <= len(sheet_name) <= 100 and  # Allow 2-char names like "VT"
                    # Sheet names should not have special chars (except spaces, dashes, Vietnamese)
                    not any(c in sheet_name for c in ['<', '>', '{', '}', '[', ']', '|', '\\', '/', '?', '*'])
                )
                
                if is_valid:
                    sheets[sheet_name] = gid
                    seen_gids.add(gid)
                    valid_count += 1
            
            if sheets:
                return sheets
            else:
                return None
            
        except Exception as e:
            print(f"Error parsing sheets from HTML: {e}")
            import traceback
            traceback.print_exc()
            return None
        
    except Exception as e:
        print(f"Error in get_sheets_from_google_spreadsheet: {e}")
        return None


def load_google_sheet(url):
    """
    Tải dữ liệu từ Google Sheet.
    Format: Cột A = Tên khách sạn, Cột B = Link
    """
    try:
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not match:
            return None, None, "URL Google Sheets không hợp lệ"
        
        spreadsheet_id = match.group(1)
        gid_match = re.search(r'[#&]gid=([0-9]+)', url)
        gid = gid_match.group(1) if gid_match else '0'

        df = None
        links_info = []
        
        # Method 0: Try Google Sheets API v4 (most reliable for public sheets)
        try:
            # Try all common gids: 0 (default), parsed gid, and some common ones
            gids_to_try = [gid, '0']
            
            for try_gid in gids_to_try:
                try:
                    api_url = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/A:Z?key=AIzaSyDummyKey&majorDimension=ROWS"
                    # Actually, we don't need API key for public sheets, try direct approach
                    api_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/gviz/tq?tqx=out:csv&gid={try_gid}"
                    
                    response = requests.get(api_url, timeout=10)
                    response.raise_for_status()
                    
                    df = pd.read_csv(StringIO(response.text), header=None)
                    
                    if len(df) > 0:
                        # Parse immediately if we got data
                        if len(df.columns) >= 2:
                            for row_idx in range(len(df)):
                                hotel_name = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else ''
                                link_value = df.iloc[row_idx, 1]
                                
                                if pd.notna(link_value):
                                    value_str = str(link_value).strip()
                                    if 'http' in value_str.lower() or 'www.' in value_str.lower():
                                        is_valid = is_booking_link(value_str)
                                        links_info.append({
                                            'row': row_idx + 1,
                                            'col': 'B',
                                            'link': value_str,
                                            'cell_value': value_str,
                                            'hotel_name': hotel_name,
                                            'is_valid': is_valid,
                                            'note': '' if is_valid else '⚠️Link không hợp lệ'
                                        })
                            
                            if links_info:
                                return df, links_info, None
                        
                        break  # Success, exit loop
                except Exception as e:
                    continue
                    
        except Exception as e:
            pass
        
        # Method 1: Try CSV export (fastest but requires public access)
        try:
            csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
            csv_response = requests.get(csv_url, timeout=10)
            csv_response.raise_for_status()
            df = pd.read_csv(StringIO(csv_response.text), header=None)
        except requests.exceptions.HTTPError as e:
            df = None
        except Exception as e:
            df = None

        # Method 2: Try HTML export for hyperlinks (works with most sheets)
        if df is None or len(df) == 0:
            try:
                html_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=html&gid={gid}"
                html_response = requests.get(html_url, timeout=10)
                html_response.raise_for_status()
                
                soup = BeautifulSoup(html_response.content, 'lxml')
                rows = soup.find_all('tr')
                
                for row_idx, tr in enumerate(rows, 1):
                    cells = tr.find_all('td')
                    if len(cells) > 1:
                        cell_a = cells[0]
                        cell_b = cells[1]
                        
                        hotel_name = cell_a.get_text(strip=True) if cell_a else ''
                        cell_text = cell_b.get_text(strip=True)
                        
                        if 'http' in cell_text.lower() or 'www.' in cell_text.lower():
                            is_valid = is_booking_link(cell_text)
                            links_info.append({
                                'row': row_idx,
                                'col': 'B',
                                'link': cell_text,
                                'cell_value': cell_text,
                                'hotel_name': hotel_name,
                                'is_valid': is_valid,
                                'note': '' if is_valid else '⚠️Link không hợp lệ'
                            })
                
                if links_info:
                    return df, links_info, None
                    
            except requests.exceptions.HTTPError as e:
                pass
            except Exception as e:
                pass

        # Method 3: Try pubhtml (for published sheets)
        if not links_info:
            try:
                pubhtml_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/pubhtml?gid={gid}"
                html_response = requests.get(pubhtml_url, timeout=10)
                html_response.raise_for_status()
                
                soup = BeautifulSoup(html_response.content, 'lxml')
                rows = soup.find_all('tr')
                
                for row_idx, tr in enumerate(rows, 1):
                    cells = tr.find_all('td')
                    if len(cells) > 1:
                        cell_a = cells[0]
                        cell_b = cells[1]
                        
                        hotel_name = cell_a.get_text(strip=True) if cell_a else ''
                        cell_text = cell_b.get_text(strip=True)
                        
                        if 'http' in cell_text.lower() or 'www.' in cell_text.lower():
                            is_valid = is_booking_link(cell_text)
                            links_info.append({
                                'row': row_idx,
                                'col': 'B',
                                'link': cell_text,
                                'cell_value': cell_text,
                                'hotel_name': hotel_name,
                                'is_valid': is_valid,
                                'note': '' if is_valid else '⚠️Link không hợp lệ'
                            })
                
                if links_info:
                    return df, links_info, None
                    
            except Exception as e:
                pass

        # Method 4: Fallback to CSV data if we got it
        if df is not None and len(df) > 0 and not links_info:
            if len(df.columns) > 1:
                for row_idx in range(len(df)):
                    hotel_name = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else ''
                    link_value = df.iloc[row_idx, 1]
                    
                    if pd.notna(link_value):
                        value_str = str(link_value).strip()
                        if 'http' in value_str.lower() or 'www.' in value_str.lower():
                            is_valid = is_booking_link(value_str)
                            links_info.append({
                                'row': row_idx + 1,
                                'col': 'B',
                                'link': value_str,
                                'cell_value': value_str,
                                'hotel_name': hotel_name,
                                'is_valid': is_valid,
                                'note': '' if is_valid else '⚠️Link không hợp lệ'
                            })
                
                if links_info:
                    return df, links_info, None
        
        # All methods failed
        if not links_info:
            error_msg = (
                "Không thể đọc dữ liệu từ Google Sheet. "
                "Vui lòng đảm bảo:\n"
                "1. Sheet được share với quyền 'Anyone with the link can view'\n"
                "2. Định dạng đúng: Cột A = Tên khách sạn, Cột B = Link\n"
                "3. Có ít nhất 1 link trong cột B"
            )
            return None, None, error_msg
        
        return df, links_info, None
        
    except Exception as e:
        error_msg = f"Lỗi khi tải Google Sheets: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        return None, None, error_msg


def load_google_sheet_all_sheets(url):
    """
    Tải dữ liệu từ TẤT CẢ các sheets trong Google Spreadsheet.
    Returns: dict {sheet_name: links_info} or (None, error_message)
    """
    try:
        # Extract spreadsheet ID
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not match:
            return None, "Không thể trích xuất spreadsheet ID từ URL"
        
        spreadsheet_id = match.group(1)
        
        # Method: Download entire spreadsheet as XLSX and parse all sheets
        export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
        
        try:
            response = requests.get(export_url, timeout=30)
            
            if response.status_code != 200:
                return None, f"Không thể tải spreadsheet (Status {response.status_code}). Vui lòng kiểm tra quyền truy cập."
            
            # Parse Excel file
            from io import BytesIO
            excel_file = BytesIO(response.content)
            xl = pd.ExcelFile(excel_file)
            
            all_sheets_data = {}
            
            for sheet_name in xl.sheet_names:
                try:
                    # Read sheet without header
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    if len(df) < 2:  # Need at least header + 1 data row
                        continue
                    
                    if len(df.columns) < 2:  # Need at least column A and B
                        continue
                    
                    # Parse links from column B (index 1)
                    links_info = []
                    
                    for row_idx in range(1, len(df)):  # Skip header row
                        hotel_name = str(df.iloc[row_idx, 0]).strip() if pd.notna(df.iloc[row_idx, 0]) else ''
                        link_value = df.iloc[row_idx, 1]
                        
                        if pd.notna(link_value):
                            value_str = str(link_value).strip()
                            
                            if 'http' in value_str.lower() or 'www.' in value_str.lower():
                                is_valid = is_booking_link(value_str)
                                links_info.append({
                                    'row': row_idx + 1,
                                    'col': 'B',
                                    'link': value_str,
                                    'cell_value': value_str,
                                    'hotel_name': hotel_name,
                                    'is_valid': is_valid,
                                    'market': sheet_name,  # Use sheet name as market
                                    'note': '' if is_valid else '⚠️Link không hợp lệ'
                                })
                    
                    if links_info:
                        all_sheets_data[sheet_name] = links_info
                
                except Exception as e:
                    continue
            
            if all_sheets_data:
                total_links = sum(len(v) for v in all_sheets_data.values())
                print(f"\n✓ Successfully loaded {len(all_sheets_data)} sheets with total {total_links} links")
                return all_sheets_data, None
            else:
                return None, "Không tìm thấy link nào trong tất cả các sheets"
        
        except requests.RequestException as e:
            return None, f"Lỗi kết nối: {str(e)}"
        
    except Exception as e:
        error_msg = f"Lỗi khi tải Google Sheets: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        return None, error_msg


def check_date_outdated(url):
    try:
        checkin_match = re.search(r'checkin=(\d{4}-\d{2}-\d{2})', url)
        checkout_match = re.search(r'checkout=(\d{4}-\d{2}-\d{2})', url)
        
        if checkin_match and checkout_match:
            checkin_date = datetime.strptime(checkin_match.group(1), '%Y-%m-%d')
            checkout_date = datetime.strptime(checkout_match.group(1), '%Y-%m-%d')
            today = datetime.now().date()
            
            if checkin_date.date() < today or checkout_date.date() < today:
                return True, f"Ngày checkin ({checkin_match.group(1)}) hoặc checkout ({checkout_match.group(1)}) đã qua"
        
        return False, None
    except Exception as e:
        return False, None


def scrape_booking_data(url, debug_mode=False, row_num=None):
    if not SELENIUM_AVAILABLE:
        return None, "Selenium chưa được cài đặt. Vui lòng chạy: pip install selenium webdriver-manager"
    
    driver = None
    forced_url = force_vnd_currency(url)

    currency_code = 'VND'
    try:
        parsed = urlparse(forced_url)
        qs = parse_qs(parsed.query)
        if qs.get('selected_currency') and qs['selected_currency'][0].strip():
            currency_code = qs['selected_currency'][0].strip()
    except Exception:
        pass

    try:
        # Sử dụng hàm get_driver để hỗ trợ cả Docker (Chrome) và Local (Edge) 
        driver = get_driver(is_headless=True)

        # Set page load timeout (longer if using proxy)
        page_load_timeout = 60  # Longer timeout for proxy
        driver.set_page_load_timeout(page_load_timeout)

        # Retry logic for page load with better error handling
        max_retries = 3
        for attempt in range(max_retries):
            try:
                # print(f"Loading URL (Attempt {attempt+1}/{max_retries}): {forced_url}")
                driver.get(forced_url)
                
                # Apply GPS spoofing and verify on first attempt (now with Vietnam proxy IP)
                if attempt == 0:
                    # Add Vietnamese cookies (enhanced with proxy Vietnam IP)
                    try:
                        vietnamese_cookies = [
                            {'name': 'booked_before', 'value': '1'},
                            {'name': 'currency', 'value': 'VND'}, 
                            {'name': 'language', 'value': 'vi'},
                            {'name': 'country', 'value': 'vn'},
                            {'name': 'market_country', 'value': 'vn'},
                            {'name': 'detected_country', 'value': 'vn'},
                            {'name': 'timezone', 'value': 'Asia/Ho_Chi_Minh'},
                            {'name': 'selected_currency', 'value': 'VND'},
                            {'name': 'lang', 'value': 'vi'}
                        ]
                        
                        for cookie in vietnamese_cookies:
                            try:
                                driver.add_cookie(cookie)
                            except:
                                pass
                    except Exception as e:
                        pass
                    
                # Additional Vietnamese location hints
                try:
                    driver.execute_script("""
                        // Override timezone
                        if (Intl && Intl.DateTimeFormat) {
                            Intl.DateTimeFormat = function(locale, options) {
                                return new (Intl.DateTimeFormat.bind(Intl.DateTimeFormat, 'vi-VN', options))();
                            };
                        }
                        
                        // Override navigator language
                        Object.defineProperty(navigator, 'language', {get: () => 'vi-VN'});
                        Object.defineProperty(navigator, 'languages', {get: () => ['vi-VN', 'vi', 'en']});
                    """) 
                except:
                    pass
                
                # Check for successful load
                WebDriverWait(driver, 30).until(  # Longer wait with proxy
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'h2.pp-header__title, h1, [data-testid="price-and-discounted-price"]'))
                )
                break # Success, exit retry loop
            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(5) # Wait before retry
                # If last attempt fails, loop will finish and code proceeds below
                # potentially leading to empty result or further error which is caught by caller

        time.sleep(random.uniform(3, 6))  # Longer wait with proxy
        
        # Force Vietnamese currency if not already set
        try:
            current_url = driver.current_url
            if 'selected_currency=VND' not in current_url:
                vnd_url = force_vnd_currency(current_url)
                if vnd_url != current_url:
                    driver.get(vnd_url)
                    time.sleep(2)
        except Exception as e:
            pass
        
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'h2.pp-header__title, h1, [data-testid="price-and-discounted-price"]'))
            )
        except:
            pass
        
        time.sleep(3)
        
        # Save HTML for debugging if debug_mode is enabled
        if debug_mode:
            try:
                # Create debug_html folder if not exists
                debug_folder = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'debug_html')
                os.makedirs(debug_folder, exist_ok=True)
                
                # Generate filename with timestamp and row number
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                row_suffix = f"_row{row_num}" if row_num else ""
                filename = f"debug_link{row_suffix}_{timestamp}.html"
                filepath = os.path.join(debug_folder, filename)
                
                # Save page source
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(driver.page_source)
            except Exception as e:
                pass
        
        result = {
            'hotel_name': None,
            'hotel_link': forced_url,
            'scrape_date': None,
            'popular_facilities': [],
            'rooms': [],
            'rating': None,
            'review_count': None
        }
        
        try:
            checkin_match = re.search(r'checkin=(\d{4}-\d{2}-\d{2})', url)
            checkout_match = re.search(r'checkout=(\d{4}-\d{2}-\d{2})', url)
            
            if checkin_match and checkout_match:
                checkin = checkin_match.group(1)
                checkout = checkout_match.group(1)
                result['scrape_date'] = f"{checkin} - {checkout}"
        except:
            pass
        
        if not result['scrape_date']:
            try:
                date_elem = driver.find_element(By.CSS_SELECTOR, '[data-testid="searchbox-dates-container"]')
                date_text = date_elem.text.strip()
                result['scrape_date'] = date_text
            except:
                try:
                    date_elem = driver.find_element(By.XPATH, '//*[contains(text(), "T7") or contains(text(), "CN")]')
                    date_text = date_elem.text.strip()
                    result['scrape_date'] = date_text
                except:
                    result['scrape_date'] = datetime.now().strftime('%Y-%m-%d')

        try:
            scripts = driver.find_elements(By.XPATH, '//script[@type="application/ld+json"]')
            for script in scripts:
                try:
                    data = json.loads(script.get_attribute('innerHTML'))
                    if isinstance(data, dict):
                        if 'name' in data and not result['hotel_name']:
                            result['hotel_name'] = data['name']
                        
                        if 'aggregateRating' in data:
                            rating_data = data['aggregateRating']
                            if 'ratingValue' in rating_data:
                                result['rating'] = str(rating_data['ratingValue'])
                            if 'reviewCount' in rating_data:
                                result['review_count'] = str(rating_data['reviewCount'])
                except Exception as e:
                    continue
        except Exception as e:
            pass

        if not result['hotel_name']:
            selectors = ['h2.pp-header__title', 'h1[data-testid="title"]', 'h1']
            for selector in selectors:
                try:
                    element = driver.find_element(By.CSS_SELECTOR, selector)
                    text = element.text.strip()
                    if text and len(text) >= 3:
                        result['hotel_name'] = text
                        break
                except:
                    continue

        # Lấy Các tiện nghi được ưa chuộng nhất
        try:
            facility_wrapper = driver.find_element(By.CSS_SELECTOR, '[data-testid="property-most-popular-facilities-wrapper"]')
            facility_items = facility_wrapper.find_elements(By.CSS_SELECTOR, 'li span.f6b6d2a959')
            for item in facility_items:
                facility_text = item.text.strip()
                if facility_text:
                    result['popular_facilities'].append(facility_text)
        except Exception as e:
            print(f"Facility extraction failed: {str(e)}")
            pass

        try:
            room_rows = driver.find_elements(By.CSS_SELECTOR, 'tr.js-rt-block-row')
            
            # If no room rows found with primary selector, try alternatives
            if not room_rows:
                # Try different selectors for Vietnamese market
                alternative_selectors = [
                    # Primary Vietnamese market selectors (from debug HTML analysis)
                    'table#hprt-table tbody tr',
                    '.hprt-table tbody tr',
                    'tr[data-block-id]',
                    'tr[data-hotel-rounded-price]', 
                    'tr.js-rt-block-row',
                    'tr.hprt-table-row',
                    'tr.e2e-hprt-table-row',
                    '.hprt-roomtype-block',
                    '.hprt-block',
                    '.hprt-price-block',
                    '.prco-wrapper',
                    '.bui-price-display',
                    
                    # Room type and content selectors
                    '.hprt-table-cell-roomtype',
                    '.hprt-roomtype-link',
                    '.hprt-roomtype-icon-link',
                    '.hp-rt-group_recommendation',
                    '.rt-bed-types',
                    '.hprt-facilities-facility',
                    '.c-occupancy-icons',
                    
                    # Price and availability selectors
                    '.js-average-per-night-price',
                    '.js-strikethrough-price',
                    '.bui-price-display__value',
                    '.hprt-nos-select',
                    '.only_x_left',
                    '.urgency_message_red',
                    '.thisRoomAvailabilityNew',
                    
                    # Vietnamese market specific content
                    '[data-block-id*="_"]',
                    '*:contains("VND")',
                    '*:contains("phòng")',
                    '*:contains("Tiết kiệm")',
                    '*:contains("Ưu Đãi")',
                    '*:contains("Được giới thiệu cho")',
                    '*:contains("Chúng tôi còn")',
                    '*:contains("Không hoàn tiền")',
                    
                    # Original alternative selectors
                    '[data-block-id]',
                    '.rt-room-type',
                    '[data-testid*="room"]',
                    '.accommodation-type',
                    '.room-recommendation',
                    '.bui-room-table__row',
                    '.hp-accommodation-block',
                    '.hp-room-types__item',
                    '.room-info-container',
                    '.roomtable',
                    '[data-room-id]',
                    '.availability-table tr',
                    '.room-option',
                    '.bicon-room + .bicon__text',
                    '.room-details-wrapper',
                    # Modern Booking.com selectors
                    '[data-testid="accommodation-option"]',
                    '[data-testid="room-option"]', 
                    '[data-testid="room-info"]',
                    '.bui-table__row',
                    '.sr-room__title-container',
                    '.room-type-block'
                ]
                
                for selector in alternative_selectors:
                    room_rows = driver.find_elements(By.CSS_SELECTOR, selector)
                    if room_rows:
                        break
            
            row_index = 0
            while row_index < len(room_rows):
                row = room_rows[row_index]
                
                try:
                    # Kiểm tra xem row có room type header không (có th.hprt-table-cell-roomtype)
                    room_type_cell = row.find_elements(By.CSS_SELECTOR, 'th.hprt-table-cell-roomtype')
                    
                    if room_type_cell:
                        # Row này có room type header, lấy thông tin chung
                        room_type_header = room_type_cell[0]
                        
                        # Lấy rowspan để biết có bao nhiêu pricing options
                        rowspan = 1
                        try:
                            rowspan_attr = room_type_header.get_attribute('rowspan')
                            if rowspan_attr:
                                rowspan = int(rowspan_attr)
                        except:
                            rowspan = 1
                        
                        # Lấy thông tin chung của room (room name, guests, bed, size)
                        room_common_data = {
                            'room_type': None,
                            'num_guests': None,
                            'bed_options': [],
                            'room_size': None,
                            'discount_percent': None
                        }
                        
                        # Lấy room name - Enhanced selectors for Vietnamese market
                        try:
                            room_name_selectors = [
                                '.hprt-roomtype-link',
                                '.hprt-roomtype-icon-link', 
                                '.hprt-roomtype-name',
                                'a[data-room-name]',
                                '.room-name',
                                '.sr-room__title',
                                'h4',
                                'h3'
                            ]
                            
                            for selector in room_name_selectors:
                                try:
                                    room_name_elem = room_type_header.find_element(By.CSS_SELECTOR, selector)
                                    room_name = room_name_elem.text.strip()
                                    if room_name and len(room_name) > 3:
                                        room_common_data['room_type'] = room_name
                                        break
                                except:
                                    continue
                        except:
                            pass
                        
                        # Lấy số lượng khách - ưu tiên tìm trong room_type_header trước
                        try:
                            num_guests_found = False
                            
                            # 1. Thử tìm text "Được giới thiệu cho X người" hoặc "Max people: X"
                            try:
                                # Tìm các thẻ h3, div, span có thể chứa thông tin này
                                candidates = room_type_header.find_elements(By.CSS_SELECTOR, '.e2e-gr-title, .hprt-occupancy-occupancy-info, .maxPersons-container, .c-occupancy-icons')
                                for cand in candidates:
                                    text = cand.get_attribute('textContent').strip()
                                    # Match "Được giới thiệu cho 2 người"
                                    match = re.search(r'(?:Được giới thiệu cho|Max people:|Số người tối đa:)\s*(\d+)', text, re.IGNORECASE)
                                    if match:
                                        room_common_data['num_guests'] = match.group(1)
                                        num_guests_found = True
                                        break
                            except:
                                pass

                            if not num_guests_found:
                                # 2. Count occupancy icons - selector: .c-occupancy-icons__adults i.bicon-occupancy
                                # Tìm trong cả row vì có thể icon nằm ở cell khác
                                occupancy_icons = row.find_elements(By.CSS_SELECTOR, '.c-occupancy-icons__adults i.bicon-occupancy')
                                if not occupancy_icons:
                                     occupancy_icons = row.find_elements(By.CSS_SELECTOR, '.c-occupancy-icons__adults .bicon.bicon-occupancy')
                                
                                if occupancy_icons:
                                    room_common_data['num_guests'] = str(len(occupancy_icons))
                                    num_guests_found = True

                            if not num_guests_found:
                                # 3. Fallback: parse from hidden text "Số người tối đa: X"
                                guests_elem = row.find_element(By.CSS_SELECTOR, '.bui-u-sr-only')
                                guests_text = guests_elem.text.strip()
                                num_match = re.search(r'Số người tối đa:\s*(\d+)', guests_text, re.IGNORECASE)
                                if not num_match:
                                    num_match = re.search(r'(\d+)\s*người', guests_text, re.IGNORECASE)
                                if num_match:
                                    room_common_data['num_guests'] = num_match.group(1)
                                    num_guests_found = True
                            
                            if not num_guests_found:
                                # 4. Another fallback specific to some layouts
                                occupancy = row.find_elements(By.CSS_SELECTOR, '.bui-icon.bui-icon--adults')
                                if occupancy:
                                    room_common_data['num_guests'] = str(len(occupancy))
                        except:
                            pass
                        
                        # Lấy thông tin giường
                        try:
                            # PRIORITY 1: Wholesalers table layout (TPI/B.V. layout)
                            bed_items = room_type_header.find_elements(By.CSS_SELECTOR, '.wholesalers_table__bed_options__text')
                            if bed_items:
                                for bed_item in bed_items:
                                    bed_text = bed_item.text.strip()
                                    # Remove SVG icon text if any
                                    bed_text = re.sub(r'<svg.*?</svg>', '', bed_text, flags=re.DOTALL)
                                    bed_text = bed_text.strip()
                                    if bed_text and 'Chọn giường' not in bed_text and 'tùy tình trạng' not in bed_text:
                                        room_common_data['bed_options'].append(bed_text)
                            
                            # PRIORITY 2: Standard rt-bed-types layout
                            if not room_common_data['bed_options']:
                                bed_items = room_type_header.find_elements(By.CSS_SELECTOR, '.rt-bed-types .rt-bed-type span')
                                if bed_items:
                                    for bed_item in bed_items:
                                        bed_text = bed_item.text.strip()
                                        if bed_text and 'Chọn giường' not in bed_text and 'tùy tình trạng' not in bed_text:
                                            room_common_data['bed_options'].append(bed_text)
                            
                            # PRIORITY 3: Direct rt-bed-type elements
                            if not room_common_data['bed_options']:
                                bed_items = room_type_header.find_elements(By.CSS_SELECTOR, '.rt-bed-type')
                                for bed_item in bed_items:
                                    bed_text = bed_item.text.strip()
                                    if bed_text and 'Chọn giường' not in bed_text and 'tùy tình trạng' not in bed_text:
                                        room_common_data['bed_options'].append(bed_text)
                            
                            # PRIORITY 4: Fallback - hprt-roomtype-bed wrapper
                            if not room_common_data['bed_options']:
                                bed_config_elems = room_type_header.find_elements(By.CSS_SELECTOR, '.hprt-roomtype-bed')
                                for bed_elem in bed_config_elems:
                                    bed_text = bed_elem.text.strip()
                                    if bed_text:
                                        lines = bed_text.split('\n')
                                        for line in lines:
                                            line = line.strip()
                                            if line and 'Chọn giường' not in line and 'tùy tình trạng' not in line:
                                                room_common_data['bed_options'].append(line)
                        except Exception as e:
                            pass
                        
                        # Lấy diện tích phòng (Enhanced for Vietnamese market)
                        try:
                            size_selectors = [
                                # From debug HTML analysis
                                '.hprt-facilities-facility[data-name-en="room size"] .bui-badge__text',
                                '.hprt-facilities-facility span:contains("m²")',
                                '.hprt-facilities-facility:contains("22 m²")',
                                
                                # Original selectors
                                '.hprt-roomtype-icon-info .bui-u-sr-only',
                                '*[contains(text(), "m²")]',
                                '*[contains(text(), "m2")]',
                                '*[contains(text(), "feet²")]',
                                '*[contains(text(), "ft²")]'
                            ]
                            
                            for selector in size_selectors:
                                try:
                                    if 'contains' in selector:
                                        size_elem = room_type_header.find_element(By.XPATH, f'.//{selector}')
                                    else:
                                        size_elem = room_type_header.find_element(By.CSS_SELECTOR, selector)
                                    
                                    size_text = size_elem.text.strip()
                                    # Match various size formats: "22 m²", "250 ft²", etc.
                                    size_match_m = re.search(r'(\d+)\s*m[²2]?', size_text, re.IGNORECASE)
                                    size_match_ft = re.search(r'(\d+)\s*(feet²?|ft²?)', size_text, re.IGNORECASE)
                                    
                                    if size_match_m:
                                        room_common_data['room_size'] = f"{size_match_m.group(1)} m²"
                                        break
                                    elif size_match_ft:
                                        room_common_data['room_size'] = f"{size_match_ft.group(1)} ft²"
                                        break
                                except:
                                    continue
                        except:
                            pass
                        
                        # Bây giờ lấy thông tin pricing cho từng option (rowspan lần)
                        for option_idx in range(rowspan):
                            pricing_row = room_rows[row_index + option_idx]
                            
                            room_data = {
                                'room_type': room_common_data['room_type'],
                                'num_guests': room_common_data['num_guests'],
                                'bed_options': room_common_data['bed_options'].copy(),
                                'room_size': room_common_data['room_size'],
                                'facilities': [],
                                'price': None,
                                'price_original': None,
                                'discount_percent': None
                            }
                            
                            try:
                                # Enhanced facilities extraction for Vietnamese market
                                facility_selectors = [
                                    # Vietnamese market specific (from debug analysis)
                                    '.hprt-table-cell-conditions .bui-list__item',
                                    '.hprt-conditions-bui .bui-list__item',
                                    '.bui-list--text .bui-list__item',
                                    '.hprt-conditions li',
                                    
                                    # Room table conditions
                                    '.hprt-table-cell-conditions li',
                                    '.hprt-conditions li',
                                    
                                    # Standard selectors
                                    '.e2e-cancellation[data-testid="cancellation-subtitle"]',
                                    '.e2e-prepayment[data-testid="prepayment-subtitle"]',
                                    '.bui-list__description',
                                    '.hprt-conditions-bui li'
                                ]
                                
                                for selector in facility_selectors:
                                    choice_elems = pricing_row.find_elements(By.CSS_SELECTOR, selector)
                                    if choice_elems:
                                        for choice in choice_elems:
                                            choice_text = choice.text.strip()
                                            # Clean up bullet points and formatting
                                            choice_text = re.sub(r'^[•\-–—]\s*', '', choice_text)
                                            choice_text = choice_text.strip()
                                            
                                            # Skip empty or very short text
                                            if choice_text and len(choice_text) > 3:
                                                # Extract meaningful parts for Vietnamese content
                                                lines = choice_text.split('\n')
                                                for line in lines:
                                                    line = line.strip()
                                                    if line and len(line) > 3:
                                                        # Skip pure price lines
                                                        if not re.search(r'^\s*\d+[\.,\d]*\s*(VND|USD|EUR)?\s*$', line):
                                                            room_data['facilities'].append(line)
                                        break  # Found facilities with this selector, move to next
                                
                                # If no facilities found, try broader search
                                if not room_data['facilities']:
                                    broader_selectors = [
                                        '.bui-list__description strong',
                                        '.hprt-table-cell-conditions strong',
                                        '.policy-title',
                                        '[data-testid="policy-title"]'
                                    ]
                                    
                                    for selector in broader_selectors:
                                        elems = pricing_row.find_elements(By.CSS_SELECTOR, selector)
                                        for elem in elems:
                                            text = elem.text.strip()
                                            if text and len(text) > 3 and text not in room_data['facilities']:
                                                room_data['facilities'].append(text)
                                                
                            except:
                                pass

                            # Lấy discount percentage (Enhanced for Vietnamese market)
                            try:
                                # Look for Vietnamese discount indicators
                                price_cell = pricing_row.find_element(By.CSS_SELECTOR, '.hprt-table-cell-price, [data-testid="price-and-discounted-price"], .hprt-price-block')
                                discount_text = price_cell.text
                                # Vietnamese patterns: "Tiết kiệm 55%", "Giảm 30%", etc.
                                discount_match = re.search(r'(?:Ti[ếe]t ki[ệe]m|Gi[aả]m)\s+(\d+)%', discount_text, re.IGNORECASE)
                                if discount_match:
                                    room_data['discount_percent'] = f"Tiết kiệm {discount_match.group(1)}%"
                                
                                # Also look for "Ưu đãi" badges
                                if not room_data['discount_percent']:
                                    discount_badges = pricing_row.find_elements(By.CSS_SELECTOR, '.bui-badge__text')
                                    for badge in discount_badges:
                                        badge_text = badge.text.strip()
                                        if 'tiết kiệm' in badge_text.lower() or 'ưu đãi' in badge_text.lower():
                                            room_data['discount_percent'] = badge_text
                                            break
                            except:
                                pass

                            # Lấy giá (Enhanced selectors for Vietnamese market)
                            price_selectors = [
                                # Primary Vietnamese market selectors (from debug analysis)
                                '.js-average-per-night-price',
                                '.bui-price-display__value .prc-no-css',
                                '.bui-price-display__value',
                                '.prco-text-nowrap-helper',
                                '.prco-f-font-heading',
                                
                                # Strikethrough (original) price selectors
                                '.js-strikethrough-price',
                                '.bui-price-display__original',
                                '.bui-f-color-destructive',
                                
                                # Other price containers
                                '[data-testid="price-and-discounted-price"] .prco-valign-middle-helper',
                                '.prco-inline-block-maker-helper',
                                '.bui_font_strong',
                                '.prco-text-color-bold',
                                '.hprt-price-block .bui-price-display',
                                '.prco-wrapper .bui-price-display',
                                
                                # Fallback selectors
                                'span[aria-hidden="true"]',
                                '.price',
                                '.rate'
                            ]
                            
                            for selector in price_selectors:
                                try:
                                    price_elem = pricing_row.find_element(By.CSS_SELECTOR, selector)
                                    price_text = price_elem.text.strip().replace('\n', ' ').replace('\xa0', ' ')
                                    if price_text and len(price_text) > 0:
                                        price_match = re.search(r'VND\s*([\d\.,]+)', price_text, re.I)
                                        if not price_match:
                                            price_match = re.search(r'([\d\.,]+)\s*VND', price_text, re.I)
                                        if not price_match:
                                            price_match = re.search(r'([\d\.,]+)', price_text)
                                        if price_match:
                                            value = price_match.group(1).replace(',', '.')
                                            room_data['price'] = f"{value} {currency_code}".strip()
                                            break
                                except:
                                    continue

                            if not room_data['price']:
                                try:
                                    per_night_elem = pricing_row.find_element(By.CSS_SELECTOR, '.js-average-per-night-price')
                                    raw_val = per_night_elem.get_attribute('data-price-per-night-raw') or per_night_elem.text
                                    if raw_val:
                                        raw_val = str(raw_val).strip()
                                        number_match = re.search(r'([\d\.,]+)', raw_val)
                                        if number_match:
                                            value = number_match.group(1).replace(',', '.')
                                            room_data['price'] = f"{value} {currency_code}".strip()
                                except:
                                    pass

                            # Lấy giá gốc
                            original_price_selectors = [
                                '.bui-price-display__original',
                                '.bui-price-display__strikethrough',
                                '[data-testid="price-and-discounted-price"] .bui-price-display__strikethrough',
                                '.prco-text-stack s'
                            ]
                            
                            for selector in original_price_selectors:
                                try:
                                    original_price = pricing_row.find_element(By.CSS_SELECTOR, selector)
                                    orig_text = original_price.text.strip().replace('\n', ' ').replace('\xa0', ' ')
                                    if orig_text and len(orig_text) > 0:
                                        orig_match = re.search(r'VND\s*([\d\.,]+)', orig_text, re.I)
                                        if not orig_match:
                                            orig_match = re.search(r'([\d\.,]+)\s*VND', orig_text, re.I)
                                        if not orig_match:
                                            orig_match = re.search(r'([\d\.,]+)', orig_text)
                                        if orig_match:
                                            value = orig_match.group(1).replace(',', '.')
                                            room_data['price_original'] = f"{value} {currency_code}".strip()
                                            break
                                except:
                                    continue

                            if not room_data['price_original']:
                                try:
                                    orig_elem = pricing_row.find_element(By.CSS_SELECTOR, '.js-strikethrough-price')
                                    raw_orig = orig_elem.get_attribute('data-strikethrough-value') or orig_elem.text
                                    if raw_orig:
                                        raw_orig = str(raw_orig).strip()
                                        number_match = re.search(r'([\d\.,]+)', raw_orig)
                                        if number_match:
                                            value = number_match.group(1).replace(',', '.')
                                            room_data['price_original'] = f"{value} {currency_code}".strip()
                                except:
                                    pass

                            # Extract Vietnamese market specific information
                            try:
                                # Room availability/scarcity information
                                scarcity_selectors = [
                                    '.only_x_left', 
                                    '.urgency_message_red',
                                    '.thisRoomAvailabilityNew',
                                    '.top_scarcity',
                                    '*:contains("Chúng tôi còn")',
                                    '*:contains("căn")'
                                ]
                                
                                for selector in scarcity_selectors:
                                    try:
                                        if 'contains' in selector:
                                            scarcity_elem = pricing_row.find_element(By.XPATH, f'.//{selector}')
                                        else:
                                            scarcity_elem = pricing_row.find_element(By.CSS_SELECTOR, selector)
                                        
                                        scarcity_text = scarcity_elem.text.strip()
                                        if scarcity_text and ('còn' in scarcity_text or 'left' in scarcity_text):
                                            room_data['availability_info'] = scarcity_text
                                            break
                                    except:
                                        continue
                            except:
                                pass
                                
                            # Extract room recommendation info  
                            try:
                                recommendation_selectors = [
                                    '.hp-rt-group_recommendation',
                                    '*:contains("Được giới thiệu cho")',
                                    '*:contains("người lớn")'
                                ]
                                
                                for selector in recommendation_selectors:
                                    try:
                                        if 'contains' in selector:
                                            rec_elem = room_type_header.find_element(By.XPATH, f'.//{selector}')
                                        else:
                                            rec_elem = room_type_header.find_element(By.CSS_SELECTOR, selector)
                                        
                                        rec_text = rec_elem.text.strip()
                                        if rec_text:
                                            room_data['recommendation'] = rec_text
                                            break
                                    except:
                                        continue
                            except:
                                pass

                            # Thêm room data vào result
                            if room_data['room_type'] and room_data['price']:
                                result['rooms'].append(room_data)
                        
                        # Tăng row_index để skip các pricing rows đã xử lý
                        row_index += rowspan
                    else:
                        # Row này không có room type header -> bỏ qua (đã được xử lý trong rowspan loop trước đó)
                        row_index += 1
                        
                except Exception as e:
                    row_index += 1
                    continue
        except Exception as e:
            print(f"Room extraction error: {str(e)}")
            pass
        
        # Luôn trả về result ngay cả khi không có phòng
        # Frontend/API sẽ quyết định xử lý như thế nào
        return result, None
        
    except Exception as e:
        return None, f"Lỗi: {str(e)}"
        
    finally:
        if driver:
            try:
                driver.quit()
            except Exception as cleanup_error:
                # If clean quit fails, try force kill
                try:
                    driver.service.process.kill()
                except:
                    pass
