from fastapi import APIRouter, HTTPException, File, UploadFile
from typing import List, Optional
from pydantic import BaseModel
from app.database.repositories import CompetitorListRepository
import openpyxl
import io

router = APIRouter(prefix="/api/competitors", tags=["competitors"])

class CompetitorData(BaseModel):
    hotel_name: Optional[str] = None
    hotel_link: Optional[str] = None
    room_type: Optional[str] = None
    num_people: Optional[int] = None
    bed_info: Optional[str] = None
    room_area: Optional[str] = None
    room_choices: Optional[str] = None
    popular_facilities: Optional[str] = None
    market: Optional[str] = None
    cluster: Optional[str] = None
    competitor_level: Optional[str] = None
    breakfast_included: Optional[str] = None
    room_group: Optional[str] = None
    level: Optional[str] = None

@router.get("")
async def get_all_competitors(limit: int = 1000, offset: int = 0):
    """Get all competitors with pagination"""
    try:
        repo = CompetitorListRepository()
        competitors = repo.get_all_competitors(limit, offset)
        total = repo.get_total_count()
        return {
            "data": competitors,
            "total": total,
            "limit": limit,
            "offset": offset
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{competitor_id}")
async def get_competitor_by_id(competitor_id: int):
    """Get competitor by ID"""
    try:
        repo = CompetitorListRepository()
        competitor = repo.get_competitor_by_id(competitor_id)
        if competitor:
            return competitor
        else:
            raise HTTPException(status_code=404, detail="Competitor not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("")
async def create_competitor(data: CompetitorData):
    """Create a new competitor"""
    try:
        repo = CompetitorListRepository()
        competitor_id = repo.create_competitor(data.dict())
        return {"id": competitor_id, "message": "Competitor created successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/{competitor_id}")
async def update_competitor(competitor_id: int, data: CompetitorData):
    """Update an existing competitor"""
    try:
        repo = CompetitorListRepository()
        success = repo.update_competitor(competitor_id, data.dict())
        if success:
            return {"message": "Competitor updated successfully"}
        else:
            raise HTTPException(status_code=404, detail="Competitor not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{competitor_id}")
async def delete_competitor(competitor_id: int):
    """Delete a competitor"""
    try:
        repo = CompetitorListRepository()
        success = repo.delete_competitor(competitor_id)
        if success:
            return {"message": "Competitor deleted successfully"}
        else:
            raise HTTPException(status_code=404, detail="Competitor not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/import")
async def import_from_excel(file: UploadFile = File(...)):
    """Import competitors from Excel file"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be Excel format")
        
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        repo = CompetitorListRepository()
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        print(f"\n=== Starting Excel Import: {file.filename} ===")
        print(f"Total rows in sheet: {sheet.max_row}")
        
        # Expected columns: A=Tên khách sạn, B=Link, C=Tên hạng phòng, D=Số người, 
        # E=Giường, F=Diện tích, G=Các lựa chọn, H=Tiện nghi, 
        # I=Market, J=Cluster, K=Level đối thủ, L=Giá bao gồm bữa sáng, M=Nhóm hạng phòng, N=Level
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Validate row is a tuple/list
                if not isinstance(row, (tuple, list)):
                    error_msg = f"Invalid row type: {type(row).__name__}"
                    errors.append({"row": row_idx, "error": error_msg})
                    print(f"Row {row_idx}: ✗ {error_msg}")
                    continue
                
                # Check if row is empty
                if not row or len(row) == 0:
                    skipped_count += 1
                    continue
                
                # Safe check for first cell
                first_cell = row[0] if len(row) > 0 else None
                second_cell = row[1] if len(row) > 1 else None
                
                # Skip rows with no hotel name AND no link
                if not first_cell and not second_cell:
                    skipped_count += 1
                    continue
                
                # Must have at least hotel name
                if not first_cell:
                    print(f"Row {row_idx}: Skipped - No hotel name")
                    skipped_count += 1
                    continue
                
                # Helper function to safely get value from row
                def get_value(idx, converter=str):
                    try:
                        if len(row) > idx and row[idx] is not None:
                            value = row[idx]
                            # Convert to string first for safety, then to target type
                            if converter == int:
                                return int(float(str(value)))
                            else:
                                return str(value).strip()
                    except Exception as conv_error:
                        print(f"    Conversion error at column {idx}: {conv_error}")
                        return None
                
                # Parse all data BEFORE attempting to save
                data = {
                    'hotel_name': get_value(0, str),
                    'hotel_link': get_value(1, str),
                    'room_type': get_value(2, str),
                    'num_people': get_value(3, int),
                    'bed_info': get_value(4, str),
                    'room_area': get_value(5, str),
                    'room_choices': get_value(6, str),
                    'popular_facilities': get_value(7, str),
                    'market': get_value(8, str),
                    'cluster': get_value(9, str),
                    'competitor_level': get_value(10, str),
                    'breakfast_included': get_value(11, str),
                    'room_group': get_value(12, str),
                    'level': get_value(13, str),
                }
                
                print(f"Row {row_idx}: Processing {data['hotel_name']}")
                
                # Only save if parsing was successful (no None for required fields)
                if not data['hotel_name']:
                    raise ValueError("Hotel name is required")
                
                # Check if competitor exists before upsert
                existing = repo.find_competitor(
                    hotel_name=data.get('hotel_name'),
                    room_type=data.get('room_type'),
                    num_people=data.get('num_people'),
                    bed_info=data.get('bed_info'),
                    room_area=data.get('room_area')
                )
                
                # Use upsert_competitor which returns the ID
                competitor_id = repo.upsert_competitor(data)
                
                if existing:
                    updated_count += 1
                    print(f"  ✓ Updated (ID: {competitor_id})")
                else:
                    created_count += 1
                    print(f"  ✓ Created (ID: {competitor_id})")
                    
            except Exception as e:
                error_msg = {"row": row_idx, "error": str(e)}
                errors.append(error_msg)
                print(f"Row {row_idx}: ✗ Error: {str(e)}")
        
        result = {
            "message": "Import completed",
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": errors
        }
        
        print(f"\n=== Import Summary ===")
        print(f"Created: {created_count}")
        print(f"Updated: {updated_count}")
        print(f"Skipped: {skipped_count}")
        print(f"Errors: {len(errors)}")
        
        return result
        
    except Exception as e:
        print(f"Import failed: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
